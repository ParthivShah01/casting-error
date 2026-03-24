import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from io import BytesIO
import re
from decimal import Decimal, ROUND_HALF_UP

# --- Excel-style rounding ---
def excel_round(x):
    try:
        return float(Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except:
        return 0

st.set_page_config(page_title="Casting Error Detector", page_icon="🧮", layout="wide")
st.title("🧮 Casting Error Detector")

# --- Checkboxes ---
col1, col2 = st.columns(2)
with col1:
    check_add = st.checkbox("Check Add / Sum / Sub formulas", value=True)
with col2:
    check_mul = st.checkbox("Check Mul / Div formulas", value=False)

uploaded_file = st.file_uploader("📂 Upload your Excel file (.xlsx)", type=["xlsx"])

if uploaded_file:
    try:
        wb_formula = openpyxl.load_workbook(uploaded_file, data_only=False)
        wb_values = openpyxl.load_workbook(uploaded_file, data_only=True)
    except Exception as e:
        st.error(f"❌ Error loading workbook: {e}")
        st.stop()

    results = []
    error_cells = {}

    for sheet_name in wb_formula.sheetnames:
        sheet_f = wb_formula[sheet_name]
        sheet_v = wb_values[sheet_name]

        for row in sheet_f.iter_rows():
            for cell in row:
                if cell.data_type == "f" and isinstance(cell.value, str):
                    formula = cell.value.strip()

                    is_mul_div = ("*" in formula or "/" in formula)

                    # --- Decide category ---
                    if is_mul_div:
                        process = check_mul
                        category = "MUL/DIV"
                    else:
                        process = check_add
                        category = "ADD/SUB"

                    if not process:
                        continue

                    try:
                        # --- SUM handling ---
                        if formula.upper().startswith("=SUM("):
                            range_part = formula.upper().replace("=SUM(", "").replace(")", "")
                            cell_range = sheet_v[range_part]

                            values = [
                                c.value for r in cell_range for c in r
                                if isinstance(c.value, (int, float))
                            ]

                            if not values:
                                continue

                            rounded_inputs = [excel_round(x) for x in values]
                            rounded_calc = excel_round(sum(rounded_inputs))

                        else:
                            # --- Arithmetic handling ---
                            expr = formula[1:].replace(" ", "")

                            refs = re.findall(r"[A-Z]+[0-9]+", expr)

                            ref_values = {}
                            for ref in refs:
                                try:
                                    v = sheet_v[ref].value
                                    ref_values[ref] = v if isinstance(v, (int, float)) else 0
                                except:
                                    ref_values[ref] = 0

                            # actual values replaced
                            eval_expr = expr
                            for ref, val in ref_values.items():
                                eval_expr = re.sub(rf"\b{ref}\b", str(val), eval_expr)

                            # rounded inputs replaced
                            rounded_expr = expr
                            for ref, val in ref_values.items():
                                rounded_expr = re.sub(rf"\b{ref}\b", str(excel_round(val)), rounded_expr)

                            rounded_calc = excel_round(eval(rounded_expr))

                        # --- Excel cell value ---
                        actual_cell_val = sheet_v[cell.coordinate].value
                        actual_cell_val = excel_round(actual_cell_val)

                        match = (rounded_calc == actual_cell_val)

                        # --- Logic based on checkbox state ---
                        if is_mul_div and not check_mul:
                            # Only comment, no checking
                            error_cells.setdefault(sheet_name, []).append(
                                (cell.coordinate, f"Rounded Value = {rounded_calc}", False)
                            )

                        elif not is_mul_div and not check_add:
                            # Only comment, no checking
                            error_cells.setdefault(sheet_name, []).append(
                                (cell.coordinate, f"Rounded Value = {rounded_calc}", False)
                            )

                        else:
                            # Perform casting check
                            if not match:
                                error_cells.setdefault(sheet_name, []).append(
                                    (cell.coordinate, f"Rounded Calculation = {rounded_calc}", True)
                                )

                        results.append({
                            "Sheet": sheet_name,
                            "Cell": cell.coordinate,
                            "Formula": formula,
                            "Rounded Calc": rounded_calc,
                            "Excel Value": actual_cell_val,
                            "Status": "✅ OK" if match else "❌ Casting Error"
                        })

                    except Exception as e:
                        st.warning(f"⚠️ Error at {cell.coordinate}: {e}")

    # ---- Display ----
    if results:
        df = pd.DataFrame(results)
        st.subheader("📊 Summary")

        def highlight_status(val):
            if "❌" in val:
                return "color: red; font-weight: bold;"
            elif "✅" in val:
                return "color: green; font-weight: bold;"
            return ""

        st.dataframe(df.style.map(highlight_status, subset=["Status"]))

        # --- Excel Highlight & Comments ---
        yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

        for sheet_name, cells in error_cells.items():
            sheet = wb_formula[sheet_name]
            for cell_ref, comment_text, highlight in cells:
                cell = sheet[cell_ref]

                if highlight:
                    cell.fill = yellow_fill

                cell.comment = Comment(comment_text, "Casting Error Detector")

        # --- Download ---
        output = BytesIO()
        wb_formula.save(output)
        output.seek(0)

        st.download_button(
            label="📥 Download Excel",
            data=output,
            file_name="CastingErrorHighlighted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    else:
        st.warning("No formulas processed.")
else:
    st.info("⬆️ Upload Excel file")
